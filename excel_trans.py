from docx import Document
import pdbc
from file_type_convert import file_type_convert
from openpyxl import load_workbook
from cutString import cutString
from clean_f import clean_f
from merge import merge
from config import Config
from utils import Utils
import threading
import math
import time

class ExcelTrans(object):
    def __init__(self) -> None:
        self.p = pdbc.PdbcConnector()
        self.file_convert = file_type_convert()
        self.cutString = cutString()
        self.clean_f = clean_f()
        self.merge = merge()
        self.utils = Utils()
    
    def sheetThread(self, wb, sheet, document_id, document_save_name, src, tgt, user_id, document_name, sheet_index):
        colThreads = []
        for i, col in enumerate(sheet.iter_cols()):
            sheet_col = str(sheet_index) + "-" + str(i)
            aThread = threading.Thread(target=self.excelTransThread,
                                       args=(wb, col, document_id, document_save_name, src, tgt, user_id, document_name, sheet_col))
            colThreads.append(aThread)

        for thread in colThreads:
            thread.start()

        for t in colThreads:
            t.join()

    def excelTransThread(self, wb, col, document_id, document_save_name, src, tgt, user_id, document_name, sheet_col):
        id_index = 0
        source_len = 0
        source_list = []
        trans_list = []
        trans_result = []
        source_content = ""
        record_src_text = ""
        record_tgt_text = ""
        try:
            for cell in col:
                if cell.value.rstrip('\n') is not None and str(cell.value).rstrip('\n').isdigit() == False:
                    clean_out = self.clean_f.clean(cell.value)
                    keep_str = []
                    if clean_out.lstrip().rstrip() == '':
                        keep_str.append(cell.value)
                        source_list.append(keep_str)
                        continue
                    cut_out = self.cutString.get_cut_word(clean_out)
                    if cut_out[0][1] == "blank":
                        cut_out.pop(0)
                    merge_out = self.merge.merge(cut_out)
                    if len(merge_out) == 0:
                        keep_str.append(cell.value)
                        source_list.append(keep_str)
                        continue
                    source_list.append(merge_out)
                    source_content = ""
                    for each in merge_out:
                        each = list(each)
                        if each[1] == src:
                            source_content = source_content + each[0] + "\n"

                    data = {}
                    data['id'] = id_index
                    data['text'] = source_content.rstrip('\n')
                    trans_list.append(data)
                    id_index += 1
                    source_len += len(source_content)
                    record_src_text += " " + source_content.strip('\n')

                    if source_len > Config.MAX_CHARACTERS:
                        result = self.utils.getNmtJson(src, tgt, trans_list, user_id)
                        #记录翻译内容到数据库
                        if result:
                            self.p.record_src_text(user_id, src, tgt, record_src_text, len(record_src_text), document_name, sheet_col, document_id)
                            record_src_text = ""
                            for i in result:
                                record_tgt_text += i['trans_text']
                            self.p.record_tgt_text(user_id, record_tgt_text, document_name, sheet_col, document_id)
                            record_tgt_text = ""  

                        #if not result:
                            #self.p.updateDocumentStatus(document_id, '3')
                            #return False
                        trans_result.extend(result)
                        source_len = 0
                        trans_list = []
            
            # self.p.record_src_text(user_id, src, tgt, record_src_text, len(record_src_text), document_name, sheet_col, document_id)
            
            if source_len != 0:
                result = self.utils.getNmtJson(src, tgt, trans_list, user_id)
                #记录翻译内容到数据库
                if result:
                    self.p.record_src_text(user_id, src, tgt, record_src_text, len(record_src_text), document_name, sheet_col, document_id)
                    record_src_text = ""
                    for i in result:
                        record_tgt_text += i['trans_text']
                    self.p.record_tgt_text(user_id, record_tgt_text, document_name, sheet_col, document_id)
                    record_tgt_text = ""  

                #if not result:
                    #self.p.updateDocumentStatus(document_id, '3')
                    #return False
                trans_result.extend(result)

            trans_result_list = []
            for each in trans_result:
                trans_result_list.extend(each['trans_text'].split("\n"))

            record_tgt_text = ""
            for cell in col:
                if cell.value.rstrip('\n') is not None and str(cell.value).rstrip('\n').isdigit() == False:
                    value = ""
                    for cut_untrans in source_list[0]:
                        if len(source_list[0]) == 1 and isinstance(cut_untrans, str):
                            value = cut_untrans
                            continue
                        elif cut_untrans[1] == src:
                            value += trans_result_list[0]
                            del (trans_result_list[0])
                        else:
                            value += cut_untrans[0]
                    cell.value = value
                    record_tgt_text += " " + value
                    del (source_list[0])
            wb.save(Config.DOC_RESULT_FOLDER + document_save_name)
            # self.p.record_tgt_text(user_id, record_tgt_text, document_name, sheet_col, document_id)
        except Exception as e:
            print(e)
            #self.p.updateDocumentStatus(document_id, '3')
            return False

    def excelTrans(self, file_path, document_save_name, src, tgt, document_id, user_id, document_name):
        # p = pdbc.PdbcConnector()
        wb = load_workbook(filename=file_path)
        # ws = wb.active

        sheetThreads = []
        for i, sheet in enumerate(wb):
            aThread = threading.Thread(target=self.sheetThread,
                                       args=(wb, sheet, document_id, document_save_name, src, tgt, user_id, document_name, i))
            sheetThreads.append(aThread)

        for thread in sheetThreads:
            thread.start()

        for t in sheetThreads:
            t.join()
        try:
            wb.save(Config.DOC_RESULT_FOLDER + document_save_name)
        except Exception as e:
            print(e)
            #self.p.updateDocumentStatus(document_id, '3')
            return False

        self.p.updateDocumentStatus(document_id, '2')
        return True
